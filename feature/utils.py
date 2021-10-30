# -*- coding: utf-8 -*-
"""
Created on 2021/10/02
@author: Leo Long
@title: 常用工具包
"""
import pandas as pd
import numpy as np
from pandas.api.types import is_numeric_dtype
from .chi2_bin import chi2_bin
from .best_ks_bin import best_ks_bin
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings


def feature_analysis(df: pd.DataFrame, X: list, y: str, bins=5, init_bins=100, init_method='qcut', init_precision=3,
                     num_fillna=-1, cate_finllna='NA', num_independ=[], cate_independ=[], lamba=0.001,
                     print_process=True,
                     report_save_path=None):
    """
    :param df: 数据集
    :param x: 待分箱特征名称
    :param y: 目标变量的名称
    :param bins: 最终的分箱数量
    :param init_bins:  初始化分箱的数量，若为空则钚进行初始化的分箱
    :param init_method: 初始化分箱的方法，仅支持 qcut 和cut两种方式
    :param init_precision: 初始化分箱时的precision
    :param num_fillna: 数值型变量的缺失值填充值
    :param cate_finllna: 类别性变量的缺失值填充值
    :param lamba: 计算iv和woe 时候的 修正值
    :param num_independ : 数值型变量的独立分箱
    :param cate_independ ：字符串型变量的独立分箱分组
    :param print_process: 是否打印分箱过程信息
    :param report_save_path: 生成报告保存的结果
    :return: type=tuple  (dti:Dataframe 基本的信息列表 , bin_config_dict:dict 变量分箱的具体信息)
    """
    df = df.copy()
    N = len(df)
    arr = []
    binConfig_dict = {}
    if report_save_path is not None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = '变量信息汇总'
        ws2 = wb.create_sheet(title='分箱信息')

    if X is None:
        X = list(df.columns.values)
        X.remove(y)

    for x in X:
        missing_index = pd.isna(df[x])
        missing_count = missing_index.sum()
        missing_rate = missing_count * 1.0 / N
        is_numeric = is_numeric_dtype(df[x])

        if is_numeric:
            fill_na = num_fillna
            df.loc[:, x] = df[x].fillna(num_fillna)
        else:
            fill_na = cate_finllna
            df.loc[:, x] = df[x].fillna(cate_finllna)

        unique_arr = pd.unique(df[x])
        unique_count = unique_arr.size

        if unique_count <= 1:
            iv = None
            method = None
            cond = None
            dti = None
        else:
            cond1 = chi2_bin(df, x, y, bins=bins, init_bins=init_bins, init_method=init_method,
                             init_precision=init_precision, print_process=print_process)
            cond2 = best_ks_bin(df, x, y, bins=bins, init_bins=init_bins, init_method=init_method,
                                init_precision=init_precision, print_process=print_process)
            if is_numeric:
                cond1.extend(num_independ)
                cond2.extend(num_independ)
                cond1 = list(set(cond1))
                cond2 = list(set(cond2))
                cond1.sort()
                cond2.sort()
            else:
                lenz = len(cate_independ)
                for k, v in cond1.items():
                    cond1[k] = v + lenz
                for k, v in cond2.items():
                    cond2[k] = v + lenz
                for i in range(len(cate_independ)):
                    cond1[cate_independ[i]] = i
                    cond2[cate_independ[i]] = i

            iv1, dti1, binned1 = make_bin(df, x, y, cond1, lamba=lamba, ret_binned=False,
                                          fill_na=fill_na)
            iv2, dti2, binned2 = make_bin(df, x, y, cond2, lamba=lamba, ret_binned=False,
                                          fill_na=fill_na)

            if iv1 >= iv2:
                iv = iv1
                dti = dti1
                method = "chi2"
                cond = cond1
            else:
                iv = iv2
                dti = dti2
                method = "best_ks"
                cond = cond2

        arr.append([x, unique_count, missing_count, missing_rate, is_numeric, iv])
        binConfig = {"dti": dti, "method": method, "cond": cond, "iv": iv}
        binConfig_dict[x] = binConfig
        if (report_save_path is not None) and (dti is not None):
            if is_numeric:
                bin_name = x + "_bin"
                dti_to_excel = dti.copy()
                dti_to_excel[bin_name] = dti_to_excel[bin_name].astype(str)
                for r in dataframe_to_rows(dti_to_excel, header=True, index=False):
                    ws2.append(r)
                ws2.append([])
            else:
                for r in dataframe_to_rows(dti, header=True, index=False):
                    ws2.append(r)
                ws2.append([])

        result = pd.DataFrame(arr, columns=["var_name", "unique_count", "missing_count", "missing_rate", "is_numeric",
                                            "iv"]).sort_values("iv", ascending=False).reset_index(drop=True)

    if report_save_path is not None:
        for r in dataframe_to_rows(result, header=True, index=False):
            ws1.append(r)
        wb.save(report_save_path)

    return result, binConfig_dict


def make_bin(df: pd.DataFrame, x: str, y: str, cond, fill_na=None, lamba=0.001, ret_binned=False,
             binned_fileds=["bin", "woe"], woe=None,precision=4):
    """
    :param df:
    :param x:
    :param y:
    :param cond:
    :param fill_na:
    :param lamba:
    :param ret_binned:
    :param binned_fileds:
    :param woe: 预设的woe 结果，格式为 { bin_num:woe值 }，woe编码时使用
    :return:
    """
    df = df[[x, y]].copy()
    x_is_numeric = is_numeric_dtype(df[x])

    if ret_binned:
        raw_index_name = df.index.name
        df = df.reset_index()
        index_name = "index" if raw_index_name is None else raw_index_name
        binned_fileds = binned_fileds.copy()

    if fill_na is not None:
        df[x] = df[x].fillna(fill_na)
    else:
        warnings.warn("强烈建议在计算分箱信息的时候，设置空值的填充值fill_na参数")

    bin_name = x + "_bin"
    if x_is_numeric and isinstance(cond, list):
        cond = list(np.round(cond,precision))
        bin_rs = pd.DataFrame({bin_name: pd.IntervalIndex.from_breaks(cond, closed="right")}) \
            .reset_index() \
            .rename({"index": "bin"}, axis=1)
        df[bin_name] = pd.cut(df[x], cond, right=True,precision=precision)
        dti = pd.crosstab(df[bin_name], df[y]) \
            .reset_index() \
            .rename({0: "negative", 1: "positive"}, axis=1)
        bin_rs = bin_rs.merge(dti, on=bin_name, how="left")
        bin_rs["positive_rate"] = bin_rs["positive"] / (bin_rs["positive"] + bin_rs["negative"])
    elif (not x_is_numeric) and isinstance(cond, dict):
        bin_rs = pd.DataFrame.from_dict(cond, orient="index") \
            .reset_index() \
            .rename({"index": bin_name, 0: "bin"}, axis=1)
        df[bin_name] = df[x]
        df = df.merge(bin_rs, on=bin_name, how="left")
        dti = pd.crosstab([df[bin_name], df["bin"]], df[y]).reset_index(). \
            rename({0: "negative", 1: "positive"}, axis=1). \
            sort_values("bin")
        bin_rs = bin_rs.merge(dti[[bin_name, "negative", "positive"]], on=bin_name, how="left")
        bin_rs["positive_rate"] = bin_rs["positive"] / (bin_rs["positive"] + bin_rs["negative"])
        mapping = bin_rs.copy()
        bin_rs = dti.groupby("bin")[["negative", "positive"]].sum().reset_index()
    else:
        raise ValueError("数字型变量cond必须为list，类别型变量cond必须为dict")
    n_t = dti["negative"].sum()
    p_t = dti["positive"].sum()

    if woe is None:
        bin_rs["woe"] = np.log(((bin_rs["positive"] / p_t) + lamba) / ((bin_rs["negative"] / n_t) + lamba))
    else:
        woe_rs = pd.DataFrame.from_dict(woe, orient="index").reset_index().rename({"index": "bin", 0: "woe"},axis=1)
        bin_rs = bin_rs.merge(woe_rs, on="bin", how="left")
    bin_rs["iv"] = ((bin_rs["positive"] / p_t) - (bin_rs["negative"] / n_t)) * np.log(
        ((bin_rs["positive"] / p_t) + lamba) / ((bin_rs["negative"] / n_t) + lamba))
    iv = bin_rs["iv"].sum()

    info = {"iv": iv, "is_numeric": x_is_numeric, "cond": cond}

    if not x_is_numeric:
        bin_rs = mapping.merge(bin_rs[["bin", "woe", "iv"]], on="bin", how="left")

    info["dti"] = bin_rs[[bin_name, "bin", "negative", "positive", "positive_rate", "woe", "iv"]]
    info["dti"].fillna(0,inplace=True)
    info["dti"].sort_values(by="bin",inplace=True)
    info["woe"] = woe
    if ret_binned:
        if x_is_numeric:
            df = df.merge(bin_rs[[bin_name, "bin", "woe"]], on=bin_name, how="left")
        else:
            df = df.merge(bin_rs[[bin_name, "woe"]], on=bin_name, how="left")
        df.set_index("index", inplace=True)
        df.index.name = raw_index_name
        info["binned"] = df[[x, "bin", "woe"]]
        # return iv,bin_rs[[bin_name,"bin","negative","positive","positive_rate","woe","iv" ]],df[[x,"bin","woe"]]
        return info["iv"], info["dti"], info["binned"]

    return info["iv"], info["dti"], None


def make_bin_old(df: pd.DataFrame, x: str, y: str, cond, fill_na=None, independ=[], lamba=0.001, ret_binned=False,
                 binned_fileds=["bin", "woe"], woe=None):
    """
    根据分箱条件，计算iv,分箱信息，
    :param df: 数据集
    :param x: 待分箱特征名称
    :param y: 目标变量的名称
    :param cond: 分箱条件，数值型变量为右边界的list，离散型变为 value:bin 这类键值对组成的dict
    :param fill_na: 空值的填充结果
    :param independ: 需独立计算分箱的变量值
    :param lamba: 计算IV和woe用到的 调整值
    :param ret_binned: 是否计算x变量对应分箱后的结果信息
    :param binned_fileds: 返回x变量对应的计算字段 目前默认支持 返回 bin 和woe
    :return:  type=tuple :  iv值 , dti分箱信息 , x映射的分箱值
    """
    df = df[[x, y]].copy()

    if ret_binned:
        raw_index_name = df.index.name
        df = df.reset_index()
        index_name = "index" if raw_index_name is None else raw_index_name
        binned_fileds = binned_fileds.copy()

    x_is_numeric = is_numeric_dtype(df[x])

    if fill_na is not None:
        df[x] = df[x].fillna(fill_na)
    else:
        warnings.warn("强烈建议在计算分箱信息的时候，设置空值的填充值")

    if x_is_numeric and isinstance(cond, list):
        bin_name = x + "_bin"
        if (independ is not None) and len(independ) > 0:
            df0 = df[df[x].isin(independ)]
            if len(df0) > 0:
                dti0 = pd.crosstab(df0[x], df0[y]).reset_index(). \
                    rename({0: "negative", 1: "positive", x: bin_name}, axis=1)
                dti0["independ"] = True
                if ret_binned:
                    df0.loc[:, bin_name] = df0[x]
            else:
                dti0 = None
            df1 = df[~df[x].isin(independ)]
        else:
            df1 = df

        df1.loc[:, bin_name] = pd.cut(df1[x], cond, right=True)
        dti1 = pd.crosstab(df1[bin_name], df1[y]).reset_index(). \
            rename({0: "negative", 1: "positive"}, axis=1)
        dti1["independ"] = False
        if dti0 is not None:
            dti = pd.concat([dti0, dti1], axis=0)
        else:
            dti = dti1
        dti = dti.reset_index(drop=True).reset_index().rename({"index": "bin"}, axis=1)

    elif (not x_is_numeric) and isinstance(cond, dict):
        if (independ is not None) and len(independ) > 0:
            min_values = min(cond.values())
            for i in range(len(independ)):
                cond[independ[-(1 + i)]] = min_values - (i + 1)
        mapping = pd.DataFrame.from_dict(cond, orient="index").reset_index()
        mapping.columns = [x, "bin"]
        mapping = mapping

        df = df.merge(mapping, on=x, how="left")
        mapping = pd.crosstab([df["bin"], df[x]], df[y]).reset_index(). \
            rename({0: "negative", 1: "positive"}, axis=1). \
            sort_values("bin")
        mapping["positive_rate"] = mapping["positive"] / (mapping["positive"] + mapping["negative"])

        dti = mapping.groupby("bin")[["negative", "positive"]].sum().reset_index()
        # dti = pd.crosstab(df["bin"], df[y]).reset_index(). \
        #     rename({0: "negative", 1: "positive"}, axis=1)
    else:
        raise ValueError("数字型变量cond必须为list，类别型变量cond必须为dict")

    n_t = dti["negative"].sum()
    p_t = dti["positive"].sum()
    dti["positive_rate"] = dti["positive"] / (dti["positive"] + dti["negative"])

    if woe is None:
        dti["woe"] = np.log(((dti["positive"] / p_t) + lamba) / ((dti["negative"] / n_t) + lamba))
    else:
        pass
    dti["iv"] = ((dti["positive"] / p_t) - (dti["negative"] / n_t)) * np.log(
        ((dti["positive"] / p_t) + lamba) / ((dti["negative"] / n_t) + lamba))

    info = {"iv": dti["iv"].sum()}

    if x_is_numeric:
        if ret_binned:
            # binned_fileds.insert(0, x)
            df_new = pd.concat([df0, df1], axis=0)
            df_new = df_new.merge(dti, on=bin_name, how="left")
            df_new.set_index("index", inplace=True, )
            df_new.index.name = raw_index_name
            info["binned"] = df_new[binned_fileds].sort_index()
        dti.rename({bin_name: x}, axis=1, inplace=True)
        dti = dti[["bin", x, "negative", "positive", "positive_rate", "woe", "iv"]]
        info["dti"] = dti
    else:
        mapping = mapping.merge(dti[["bin", "woe", "iv"]], on="bin", how='left')
        info["dti"] = mapping
        if ret_binned:
            # binned_fileds.insert(0, x)
            df_new = df.merge(dti, on="bin", how="left")
            df_new.set_index("index", inplace=True)
            df_new.index.name = raw_index_name
            info["binned"] = df_new[binned_fileds].sort_index()

    if ret_binned:
        return info["iv"], info["dti"], info["binned"]
    else:
        return info["iv"], info["dti"], None


def calc_bin_cond(df, x, y=None, method="chi2", fill_na=None, bins=5, init_bins=100,
                  init_method='qcut', init_precision=3, print_process=True):
    """
    在外部接口封装了chi2_bin 与  best_ks_bin
    :param df: 数据集
    :param x: 待分箱特征名称
    :param y: 目标变量的名称
    :param method: chi2 和 best_ks
    :param bins: 最终的分箱数量
    :param init_bins:  初始化分箱的数量，若为空则钚进行初始化的分箱
    :param init_method: 初始化分箱的方法，仅支持 qcut 和cut两种方式
    :param init_precision: 初始化分箱时的precision
    :param print_process: 是否答应分箱的过程信息
    :return: 数值型变量为右边界的list，离散型变为 value:bin 这类键值对组成的dict
    :return: type = list or dict 数值型变量为右边界的list，离散型变为 value:bin 这类键值对组成的dict
    """
    df = df[[x, y]].copy()
    if fill_na is not None:
        df[x] = df[x].fillna(fill_na)
    else:
        raise Warning("计算分箱时，若不设置fill_na，空值会被过滤")
    if method == "chi2":
        return chi2_bin(df, x, y, bins=bins, init_bins=init_bins, init_method=init_method,
                        init_precision=init_precision, print_process=print_process)
    elif method == "best_ks":
        return best_ks_bin(df, x, y, bins=bins, init_bins=init_bins, init_method=init_method,
                           init_precision=init_precision, print_process=print_process)
    else:
        raise ValueError("计算分箱的方式暂时只支持 chi2 和 best_ks")


def feature_analysis_old(df: pd.DataFrame, X: list, y: str, bins=5, init_bins=100, init_method='qcut', init_precision=3,
                         num_fillna=-1, cate_finllna='NA', lamba=0.001, independ=[], print_process=True,
                         report_save_path=None):
    """
    :param df: 数据集
    :param x: 待分箱特征名称
    :param y: 目标变量的名称
    :param bins: 最终的分箱数量
    :param init_bins:  初始化分箱的数量，若为空则钚进行初始化的分箱
    :param init_method: 初始化分箱的方法，仅支持 qcut 和cut两种方式
    :param init_precision: 初始化分箱时的precision
    :param num_fillna: 数值型变量的缺失值填充值
    :param cate_finllna: 类别性变量的缺失值填充值
    :param lamba: 计算iv和woe 时候的 修正值
    :param independ: 需要独立分箱的值
    :param print_process: 是否打印分箱过程信息
    :param report_save_path: 生成报告保存的结果
    :return: type=tuple  (dti:Dataframe 基本的信息列表 , bin_config_dict:dict 变量分箱的具体信息)
    """
    df = df.copy()
    N = len(df)
    arr = []
    binConfig_dict = {}
    if report_save_path is not None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = '变量信息汇总'
        ws2 = wb.create_sheet(title='分箱信息')

    if X is None:
        X = list(df.columns.values)
        X.remove(y)

    for x in X:
        missing_index = pd.isna(df[x])
        missing_count = missing_index.sum()
        missing_rate = missing_count * 1.0 / N
        is_numeric = is_numeric_dtype(df[x])

        if is_numeric:
            fill_na = num_fillna
            df.loc[:, x] = df[x].fillna(num_fillna)
        else:
            fill_na = cate_finllna
            df.loc[:, x] = df[x].fillna(cate_finllna)

        unique_arr = pd.unique(df[x])
        unique_count = unique_arr.size

        if unique_count <= 1:
            iv = None
            method = None
            cond = None
            dti = None
        else:
            cond1 = chi2_bin(df, x, y, bins=bins, init_bins=init_bins, init_method=init_method,
                             init_precision=init_precision, print_process=print_process)
            iv1, dti1, binned1 = make_bin_old(df, x, y, cond1, lamba=lamba, independ=independ, ret_binned=True,
                                              fill_na=fill_na)

            cond2 = best_ks_bin(df, x, y, bins=bins, init_bins=init_bins, init_method=init_method,
                                init_precision=init_precision, print_process=print_process)
            iv2, dti2, binned2 = make_bin_old(df, x, y, cond2, lamba=lamba, independ=independ, ret_binned=True,
                                              fill_na=fill_na)

            if iv1 >= iv2:
                iv = iv1
                dti = dti1
                method = "chi2"
                cond = cond1
            else:
                iv = iv2
                dti = dti2
                method = "best_ks"
                cond = cond2

        arr.append([x, unique_count, missing_count, missing_rate, is_numeric, iv])
        binConfig = {"dti": dti, "method": method, "cond": cond, "independ": independ,
                     "iv": iv}
        binConfig_dict[x] = binConfig
        if (report_save_path is not None) and (dti is not None):
            if is_numeric:
                dti_to_excel = dti.copy()
                dti_to_excel[x] = dti_to_excel[x].astype(str)
                for r in dataframe_to_rows(dti_to_excel, header=True, index=False):
                    ws2.append(r)
                ws2.append([])
            else:
                for r in dataframe_to_rows(dti, header=True, index=False):
                    ws2.append(r)
                ws2.append([])

        result = pd.DataFrame(arr, columns=["var_name", "unique_count", "missing_count", "missing_rate", "is_numeric",
                                            "iv"]).sort_values("iv", ascending=False).reset_index(drop=True)

    if report_save_path is not None:
        for r in dataframe_to_rows(result, header=True, index=False):
            ws1.append(r)
        wb.save(report_save_path)

    return result, binConfig_dict
